import tika
from tika import parser
import os
import openpyxl
from openpyxl import Workbook

def dictMachine(item):
    split = item.split(":")
    key = split[0]
    value = split[1].strip()
    d = {}
    d[key] = value
    return d

def tupleMachine(row):
    global stTotal
    global workOrderNum
    global otTotal
    global jobNumber
    for index in range(len(row)):
        for key in row[index]:
            if "ST-WK-TOTAL" in key:
                stTotal = row[index].get(key)
            if "WORK ORDER #" in key:
                workOrderNum = row[index].get(key)
            if "COST CODE" in key:
                costCode = row[index].get(key)
            if "OT-WK-TOTAL" in key:
                otTotal = row[index].get(key)
            if "JOB #" in key:
                jobNumber = row[index].get(key)
    return stTotal, otTotal, workOrderNum, str(costCode)[1:-1].strip(), jobNumber # Have to remove brackets & spaces from cost code

timecardDirectory = os.getcwd()
print('Takes hours off timecards and creates an excel spreadsheet.\nThis should be placed in the same folder as all timecards to be scraped.\nThe spreadsheet will be created in the same folder.')
while True:
    try:
        response = input('\nProcess all timecards contained in `' + timecardDirectory + '`? (Y \ n): ')
        if response == 'Y':
            print('\nStarting...')
            break
        else:
            print('\nMove this program to the correct folder, then rerun it.')
    except:
        pass

wb = Workbook()
ws = wb.active
ws.append(["DATE", "NAME", "EMPLOYEE #", "JOB #", "DEVELOPMENT", "BIM COORDINATION", "EQUIPMENT", "UNDERGROUND", "WALLS", "OVERHEAD", "LIGHTING", "FINISHES", "WIRE", "UNCATEGORIZED", "VACATION"])
next_row = 2
timecardCount = 0
notProcessedList = []

for file in os.listdir(timecardDirectory):
    if file.endswith('.pdf') and 'PROCESSED' not in file:
        row1 = []
        row2 = []
        row3 = []
        row4 = []
        row5 = []
        row6 = []
        row7 = []
        row8 = []
        row9 = []
        row10 = []
        timecard = parser.from_file(file)
        timecardContent = timecard['content']
        if timecardContent:
            split = timecardContent.split('\n')
            split.sort()

            for item in split:
                item = item.strip('\t')
                if item:
                    if item.startswith('EmployeeName'):
                        split = item.split(': ')
                        employeeName = split[1]
                    if item.startswith('EmployeeNumber:'):
                        split = item.split(' ')
                        employeeNumber = split[1]
                    if item.startswith('WeekEnding:'):
                        split = item.split(' ')
                        weekEnding = split[1]
                    if item.startswith('11-VC-WK-TOTAL:'):
                        split = item.split(' ')
                        totalVacation = split[1]
                    if item.startswith('1-'):
                        row1.append(dictMachine(item))
                    if item.startswith('2-'):
                        row2.append(dictMachine(item))
                    if item.startswith('3-'):
                        row3.append(dictMachine(item))
                    if item.startswith('4-'):
                        row4.append(dictMachine(item))
                    if item.startswith('5-'):
                        row5.append(dictMachine(item))
                    if item.startswith('6-'):
                        row6.append(dictMachine(item))
                    if item.startswith('7-'):
                        row7.append(dictMachine(item))
                    if item.startswith('8-'):
                        row8.append(dictMachine(item))
                    if item.startswith('9-'):
                        row9.append(dictMachine(item))
                    if item.startswith('10-'):
                        row10.append(dictMachine(item))

            # STTOTAL, OTTOTAL, WORKORDERNUM, COSTCODE, JOBNUM
            row1Tuple = tupleMachine(row1)
            row2Tuple = tupleMachine(row2)
            row3Tuple = tupleMachine(row3)
            row4Tuple = tupleMachine(row4)
            row5Tuple = tupleMachine(row5)
            row6Tuple = tupleMachine(row6)
            row7Tuple = tupleMachine(row7)
            row8Tuple = tupleMachine(row8)
            row9Tuple = tupleMachine(row9)
            row10Tuple = tupleMachine(row10)

            rowTupleList = [row1Tuple, row2Tuple, row3Tuple, row4Tuple, row5Tuple, row6Tuple, row7Tuple, row8Tuple, row9Tuple, row10Tuple]

            for rowTuple in rowTupleList:
                if rowTuple[0] != '0' or rowTuple[1] != '0':
                    totalHours = float(rowTuple[0]) + float(rowTuple[1])
                    employeeNumber = int(employeeNumber)
                    if rowTuple[4] == '' or rowTuple[4] == None:
                        jobNumber = '0000'
                    else:
                        jobNumber = rowTuple[4]

                    if rowTuple[3] == '1':
                        ws.cell(column = 1, row = next_row, value = weekEnding)
                        ws.cell(column = 2, row = next_row, value = employeeName)
                        ws.cell(column = 3, row = next_row, value = employeeNumber)
                        ws.cell(column = 4, row = next_row, value = jobNumber)
                        ws.cell(column = 5, row = next_row, value = totalHours)
                        next_row += 1
                    if rowTuple[3] == '11':
                        ws.cell(column = 1, row = next_row, value = weekEnding)
                        ws.cell(column = 2, row = next_row, value = employeeName)
                        ws.cell(column = 3, row = next_row, value = employeeNumber)
                        ws.cell(column = 4, row = next_row, value = jobNumber)
                        ws.cell(column = 6, row = next_row, value = totalHours)
                        next_row += 1
                    if rowTuple[3] == '21':
                        ws.cell(column = 1, row = next_row, value = weekEnding)
                        ws.cell(column = 2, row = next_row, value = employeeName)
                        ws.cell(column = 3, row = next_row, value = employeeNumber)
                        ws.cell(column = 4, row = next_row, value = jobNumber)
                        ws.cell(column = 7, row = next_row, value = totalHours)
                        next_row += 1
                    if rowTuple[3] == '31':
                        ws.cell(column = 1, row = next_row, value = weekEnding)
                        ws.cell(column = 2, row = next_row, value = employeeName)
                        ws.cell(column = 3, row = next_row, value = employeeNumber)
                        ws.cell(column = 4, row = next_row, value = jobNumber)
                        ws.cell(column = 8, row = next_row, value = totalHours)
                        next_row += 1
                    if rowTuple[3] == '41':
                        ws.cell(column = 1, row = next_row, value = weekEnding)
                        ws.cell(column = 2, row = next_row, value = employeeName)
                        ws.cell(column = 3, row = next_row, value = employeeNumber)
                        ws.cell(column = 4, row = next_row, value = jobNumber)
                        ws.cell(column = 9, row = next_row, value = totalHours)
                        next_row += 1
                    if rowTuple[3] == '51':
                        ws.cell(column = 1, row = next_row, value = weekEnding)
                        ws.cell(column = 2, row = next_row, value = employeeName)
                        ws.cell(column = 3, row = next_row, value = employeeNumber)
                        ws.cell(column = 4, row = next_row, value = jobNumber)
                        ws.cell(column = 10, row = next_row, value = totalHours)
                        next_row += 1
                    if rowTuple[3] == '61':
                        ws.cell(column = 1, row = next_row, value = weekEnding)
                        ws.cell(column = 2, row = next_row, value = employeeName)
                        ws.cell(column = 3, row = next_row, value = employeeNumber)
                        ws.cell(column = 4, row = next_row, value = jobNumber)
                        ws.cell(column = 11, row = next_row, value = totalHours)
                        next_row += 1
                    if rowTuple[3] == '71':
                        ws.cell(column = 1, row = next_row, value = weekEnding)
                        ws.cell(column = 2, row = next_row, value = employeeName)
                        ws.cell(column = 3, row = next_row, value = employeeNumber)
                        ws.cell(column = 4, row = next_row, value = jobNumber)
                        ws.cell(column = 12, row = next_row, value = totalHours)
                        next_row += 1
                    if rowTuple[3] == '81':
                        ws.cell(column = 1, row = next_row, value = weekEnding)
                        ws.cell(column = 2, row = next_row, value = employeeName)
                        ws.cell(column = 3, row = next_row, value = employeeNumber)
                        ws.cell(column = 4, row = next_row, value = jobNumber)
                        ws.cell(column = 13, row = next_row, value = totalHours)
                        next_row += 1
                    if rowTuple[3] == '':
                        ws.cell(column = 1, row = next_row, value = weekEnding)
                        ws.cell(column = 2, row = next_row, value = employeeName)
                        ws.cell(column = 3, row = next_row, value = employeeNumber)
                        ws.cell(column = 4, row = next_row, value = jobNumber)
                        ws.cell(column = 14, row = next_row, value = totalHours)
                        next_row += 1


            if totalVacation != '0':
                ws.cell(column = 1, row = next_row, value = weekEnding)
                ws.cell(column = 2, row = next_row, value = employeeName)
                ws.cell(column = 3, row = next_row, value = employeeNumber)
                ws.cell(column = 4, row = next_row, value = '0000')
                ws.cell(column = 15, row = next_row, value = float(totalVacation))
                next_row += 1
        
            timecardCount += 1
            print("PROCESSED " + file)
        else:
            notProcessedList.append(file)


print("\n\n" + str(timecardCount) + " timecards processed")
if len(notProcessedList) != 0:
    print("\nThe following timecards could not be processed:")
    for file in notProcessedList:
        print(file)
print("\nSaving new workbook `Timecard_Spreadsheet.xlsx`...")
wb.save('Timecard_Spreadsheet.xlsx')
print("\n--DONE--\n")
